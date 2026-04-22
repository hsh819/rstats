"""Excel I/O helpers — 保留多行 JSON 列，避免 pandas to_excel 默认行高压扁。"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


DEFAULT_WIDTHS = {
    "编号": 10, "问题类型": 14, "问题": 40,
    "SQL 查询语句": 60, "SQL查询语法": 60,
    "图形格式": 10, "回答": 80, "图表": 36, "研报截图": 36,
}


def write_task_results(
    path: Path,
    rows: Iterable[dict],
    columns: list[str],
    sheet_name: str = "Sheet1",
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        cells = [row.get(c, "") for c in columns]
        ws.append([_serialize(v) for v in cells])
        for col_idx, v in enumerate(cells, start=1):
            ws.cell(row=ws.max_row, column=col_idx).alignment = wrap
    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = DEFAULT_WIDTHS.get(col, 20)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_task_results_with_images(
    path: Path,
    rows: Iterable[dict],
    columns: list[str],
    image_col: str,
    image_dir: Path,
    sheet_name: str = "Sheet1",
    image_height_px: int = 150,
    image_width_px: int = 240,
    extra_image_cols: Optional[dict[str, str]] = None,
) -> None:
    """把图表文件 inline 嵌入到指定列。

    Parameters
    ----------
    image_col       列名，单元格内容为图片文件名（相对 image_dir）；空串则跳过。
    extra_image_cols {列名: 行 dict 中的字段名}，用于嵌入研报截图等额外图片列。
    """
    rows_list = list(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    wrap = Alignment(wrap_text=True, vertical="top")

    image_col_idx = columns.index(image_col) + 1 if image_col in columns else None
    extra_map = extra_image_cols or {}
    extra_indices = {col: columns.index(col) + 1 for col in extra_map if col in columns}

    for row in rows_list:
        cells = [row.get(c, "") for c in columns]
        # 图片列单元格写入文件名（便于用户追溯）
        ws.append([_serialize(v) for v in cells])
        r_idx = ws.max_row
        for col_idx, _ in enumerate(cells, start=1):
            ws.cell(row=r_idx, column=col_idx).alignment = wrap

        # 设置行高便于显示图片
        if image_col_idx or extra_indices:
            ws.row_dimensions[r_idx].height = 120

        # 插入主图
        if image_col_idx:
            fname = str(row.get(image_col, "") or "").strip()
            if fname:
                img_path = image_dir / fname
                if img_path.exists():
                    _insert_image(ws, img_path, r_idx, image_col_idx, image_width_px, image_height_px)

        # 插入额外图（如 paper_image）
        for col, field_name in extra_map.items():
            if col not in extra_indices:
                continue
            fname = str(row.get(field_name, "") or "").strip()
            if not fname:
                continue
            img_path = image_dir / fname if not Path(fname).is_absolute() else Path(fname)
            if img_path.exists():
                _insert_image(ws, img_path, r_idx, extra_indices[col], image_width_px, image_height_px)

    for col_idx, col in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = DEFAULT_WIDTHS.get(col, 20)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _insert_image(ws, img_path: Path, row_idx: int, col_idx: int, width_px: int, height_px: int) -> None:
    try:
        img = XLImage(str(img_path))
        img.width = width_px
        img.height = height_px
        anchor = f"{get_column_letter(col_idx)}{row_idx}"
        ws.add_image(img, anchor)
    except Exception as e:
        print(f"[excel] insert image fail {img_path}: {e}")


def _serialize(v):
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False, indent=2)
    return v if v is not None else ""
